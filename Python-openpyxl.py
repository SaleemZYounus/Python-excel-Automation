import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']


print(sheet.max_row)
print(sheet.max_column)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * .9
    corrected_cell = sheet.cell(row, 4) #creates new row
    corrected_cell.value = corrected_price

sheet["d1"].value = "new price" #adds title to D1

values = Reference(sheet, 
        min_row = 2, 
        max_row = sheet.max_row,
        min_col = 4,
        max_col = 4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')


wb.save('transactions2.xlsx')
